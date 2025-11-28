# UDISE Data Generator - Enhanced UI v3.0
import streamlit as st
import pandas as pd
import numpy as np
import os
import re
import requests
from io import BytesIO
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE CONFIG & CUSTOM STYLING
# ═══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="UDISE Data Generator",
    page_icon="🏫",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better UI - works in both light and dark modes
st.markdown("""
<style>
    /* Root variables for theming */
    :root {
        --primary-color: #6366f1;
        --primary-light: #818cf8;
        --primary-dark: #4f46e5;
        --success-color: #10b981;
        --warning-color: #f59e0b;
        --error-color: #ef4444;
    }

    /* Main container styling */
    .main .block-container {
        padding-top: 1rem;
        padding-bottom: 2rem;
        max-width: 1200px;
    }

    /* Header styling - gradient that works in both themes */
    .main-header {
        background: linear-gradient(135deg, #6366f1 0%, #8b5cf6 50%, #a855f7 100%);
        padding: 2rem 2.5rem;
        border-radius: 16px;
        margin-bottom: 1.5rem;
        color: white;
        box-shadow: 0 10px 40px rgba(99, 102, 241, 0.3);
        text-align: center;
    }

    .main-header h1 {
        margin: 0;
        font-size: 2.2rem;
        font-weight: 800;
        letter-spacing: -0.5px;
    }

    .main-header p {
        margin: 0.75rem 0 0 0;
        opacity: 0.95;
        font-size: 1.1rem;
    }

    /* Welcome card */
    .welcome-card {
        background: linear-gradient(135deg, #f0f9ff 0%, #e0f2fe 100%);
        border: 2px solid #7dd3fc;
        border-radius: 16px;
        padding: 1.5rem 2rem;
        margin: 1rem 0;
    }

    [data-theme="dark"] .welcome-card,
    .stApp[data-theme="dark"] .welcome-card {
        background: linear-gradient(135deg, #1e293b 0%, #334155 100%);
        border-color: #6366f1;
    }

    .welcome-card h3 {
        color: #0369a1;
        margin: 0 0 0.75rem 0;
        font-size: 1.2rem;
    }

    .welcome-card p, .welcome-card li {
        color: #475569;
        margin: 0.25rem 0;
        font-size: 0.95rem;
    }

    /* Feature cards */
    .feature-card {
        background: linear-gradient(145deg, #ffffff 0%, #f8fafc 100%);
        border: 1px solid #e2e8f0;
        border-radius: 12px;
        padding: 1.25rem;
        text-align: center;
        transition: all 0.3s ease;
        height: 100%;
        box-shadow: 0 2px 8px rgba(0,0,0,0.04);
    }

    .feature-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 25px rgba(99, 102, 241, 0.15);
        border-color: #6366f1;
    }

    .feature-icon {
        font-size: 2rem;
        margin-bottom: 0.5rem;
    }

    .feature-card h4 {
        color: #1e293b;
        margin: 0.5rem 0;
        font-size: 1rem;
    }

    .feature-card p {
        color: #64748b;
        margin: 0;
        font-size: 0.85rem;
    }

    /* Step indicator cards */
    .step-card {
        background: linear-gradient(145deg, #ffffff 0%, #f1f5f9 100%);
        border: 1px solid #e2e8f0;
        border-radius: 12px;
        padding: 1rem 1.25rem;
        margin-bottom: 0.75rem;
        border-left: 4px solid #6366f1;
        transition: all 0.2s ease;
    }

    .step-card:hover {
        border-left-color: #a855f7;
        box-shadow: 0 4px 12px rgba(0,0,0,0.05);
    }

    .step-card.active {
        border-left-color: #10b981;
        background: linear-gradient(145deg, #f0fdf4 0%, #dcfce7 100%);
    }

    .step-card h3 {
        color: #1e293b;
        margin: 0 0 0.25rem 0;
        font-size: 1rem;
    }

    .step-card p {
        color: #64748b;
        margin: 0;
        font-size: 0.85rem;
    }

    /* Stats cards */
    .stat-card {
        background: linear-gradient(145deg, #ffffff 0%, #f8fafc 100%);
        border: 1px solid #e2e8f0;
        border-radius: 12px;
        padding: 1rem;
        text-align: center;
        box-shadow: 0 2px 8px rgba(0,0,0,0.04);
    }

    .stat-value {
        font-size: 1.75rem;
        font-weight: 700;
        background: linear-gradient(135deg, #6366f1, #a855f7);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
    }

    .stat-label {
        font-size: 0.8rem;
        color: #64748b;
        margin-top: 0.25rem;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }

    /* Info/Success/Warning boxes */
    .info-box {
        background: linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%);
        border: 1px solid #93c5fd;
        border-radius: 10px;
        padding: 1rem 1.25rem;
        margin: 0.75rem 0;
        color: #1e40af;
    }

    .success-box {
        background: linear-gradient(135deg, #f0fdf4 0%, #dcfce7 100%);
        border: 1px solid #86efac;
        border-radius: 10px;
        padding: 1rem 1.25rem;
        margin: 0.75rem 0;
        color: #166534;
    }

    .warning-box {
        background: linear-gradient(135deg, #fffbeb 0%, #fef3c7 100%);
        border: 1px solid #fcd34d;
        border-radius: 10px;
        padding: 1rem 1.25rem;
        margin: 0.75rem 0;
        color: #92400e;
    }

    /* Mode toggle */
    .mode-toggle {
        display: inline-flex;
        background: #f1f5f9;
        border-radius: 25px;
        padding: 4px;
        gap: 4px;
    }

    .mode-btn {
        padding: 8px 16px;
        border-radius: 20px;
        border: none;
        cursor: pointer;
        font-weight: 600;
        transition: all 0.2s;
    }

    .mode-btn.active {
        background: #6366f1;
        color: white;
    }

    /* Button styling */
    .stButton > button {
        border-radius: 10px;
        font-weight: 600;
        transition: all 0.3s ease;
        border: none;
    }

    .stButton > button:hover {
        transform: translateY(-1px);
        box-shadow: 0 6px 20px rgba(99, 102, 241, 0.25);
    }

    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #6366f1 0%, #8b5cf6 100%);
    }

    /* Tab styling */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background: linear-gradient(135deg, #f8fafc 0%, #f1f5f9 100%);
        padding: 8px;
        border-radius: 12px;
        border: 1px solid #e2e8f0;
    }

    .stTabs [data-baseweb="tab"] {
        border-radius: 8px;
        padding: 0.6rem 1.2rem;
        font-weight: 600;
        color: #64748b;
    }

    .stTabs [data-baseweb="tab"][aria-selected="true"] {
        background: linear-gradient(135deg, #6366f1 0%, #8b5cf6 100%);
        color: white;
    }

    /* Sidebar styling */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #f8fafc 0%, #f1f5f9 100%);
    }

    section[data-testid="stSidebar"] .block-container {
        padding-top: 1rem;
    }

    /* Field chip */
    .field-chip {
        display: inline-block;
        background: linear-gradient(135deg, #e0e7ff 0%, #c7d2fe 100%);
        color: #4338ca;
        padding: 0.3rem 0.85rem;
        border-radius: 20px;
        font-size: 0.85rem;
        margin: 0.2rem;
        font-weight: 500;
    }

    /* Divider */
    .divider {
        height: 2px;
        background: linear-gradient(90deg, transparent, #e2e8f0, transparent);
        margin: 1.5rem 0;
    }

    /* Quick action buttons */
    .quick-action {
        background: linear-gradient(145deg, #ffffff 0%, #f8fafc 100%);
        border: 2px solid #e2e8f0;
        border-radius: 12px;
        padding: 1.25rem;
        text-align: center;
        cursor: pointer;
        transition: all 0.3s ease;
    }

    .quick-action:hover {
        border-color: #6366f1;
        box-shadow: 0 8px 25px rgba(99, 102, 241, 0.15);
        transform: translateY(-2px);
    }

    /* Help tooltip */
    .help-tip {
        display: inline-block;
        background: #6366f1;
        color: white;
        width: 18px;
        height: 18px;
        border-radius: 50%;
        text-align: center;
        font-size: 12px;
        line-height: 18px;
        cursor: help;
        margin-left: 4px;
    }

    /* Pivot table styling */
    .pivot-section {
        background: linear-gradient(145deg, #fdf4ff 0%, #fae8ff 100%);
        border: 1px solid #e879f9;
        border-radius: 12px;
        padding: 1.5rem;
        margin: 1rem 0;
    }

    /* Demo badge */
    .demo-badge {
        background: linear-gradient(135deg, #fef3c7 0%, #fde68a 100%);
        color: #92400e;
        padding: 0.25rem 0.75rem;
        border-radius: 20px;
        font-size: 0.75rem;
        font-weight: 600;
        display: inline-block;
        margin-left: 8px;
    }

    /* Hide default streamlit elements */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}

    /* Multiselect styling */
    .stMultiSelect [data-baseweb="tag"] {
        background: linear-gradient(135deg, #6366f1 0%, #8b5cf6 100%);
    }

    /* Expander styling */
    .streamlit-expanderHeader {
        font-weight: 600;
        font-size: 1rem;
        color: #1e293b;
    }
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def safe_numeric_sum(df: pd.DataFrame, cols: List[str]) -> pd.Series:
    """Sum columns coercing missing / non-numeric to 0."""
    series_list = []
    for c in cols:
        if c in df.columns:
            series_list.append(pd.to_numeric(df[c], errors="coerce").fillna(0))
        else:
            series_list.append(pd.Series([0] * len(df), index=df.index))
    if not series_list:
        return pd.Series([0] * len(df), index=df.index)
    return sum(series_list)

def to_excel_bytes_styled(df: pd.DataFrame, header_fill_color="6366f1") -> bytes:
    """Write df to an excel file in-memory with header styling."""
    wb = Workbook()
    ws = wb.active
    ws.title = "UDISE_Extract"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value) if cell.value else ""
            except:
                val = ""
            max_length = max(max_length, len(val))
        ws.column_dimensions[column].width = min(50, max(10, max_length + 2))

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()

def find_column(df, candidates):
    """Find first matching column from candidates list."""
    for c in candidates:
        if c in df.columns:
            return c
    return None

def get_numeric_columns(df):
    """Get list of columns that can be converted to numeric."""
    numeric_cols = []
    for c in df.columns:
        tmp = pd.to_numeric(df[c], errors="coerce")
        if not tmp.isnull().all():
            numeric_cols.append(c)
    return numeric_cols

def build_class_totals(target_df):
    """Create Class1_Total ... Class12_Total in the given dataframe."""
    for col in target_df.columns:
        if re.match(r"(?i)^Class\d+_(Boys|Girls|Transgen)$", col):
            target_df[col] = pd.to_numeric(target_df[col], errors="coerce").fillna(0)

    created = []
    for i in range(1, 13):
        members = [f"Class{i}_Boys", f"Class{i}_Girls", f"Class{i}_Transgen"]
        target_df[f"Class{i}_Total"] = safe_numeric_sum(target_df, members)
        created.append(f"Class{i}_Total")
    return created

def build_enrollment_presets(target_df):
    """Create Enrollment aggregations."""
    for i in range(1, 13):
        if f"Class{i}_Total" not in target_df.columns:
            members = [f"Class{i}_Boys", f"Class{i}_Girls", f"Class{i}_Transgen"]
            target_df[f"Class{i}_Total"] = safe_numeric_sum(target_df, members)

    target_df["Enrollment_1_5"] = safe_numeric_sum(target_df, [f"Class{i}_Total" for i in range(1, 6)])
    target_df["Enrollment_6_8"] = safe_numeric_sum(target_df, [f"Class{i}_Total" for i in range(6, 9)])
    target_df["Enrollment_9_10"] = safe_numeric_sum(target_df, [f"Class{i}_Total" for i in range(9, 11)])
    target_df["Enrollment_11_12"] = safe_numeric_sum(target_df, [f"Class{i}_Total" for i in range(11, 13)])
    target_df["Total_Enrollment"] = safe_numeric_sum(target_df, [f"Class{i}_Total" for i in range(1, 13)])

    return ["Enrollment_1_5", "Enrollment_6_8", "Enrollment_9_10", "Enrollment_11_12", "Total_Enrollment"]

def create_demo_data():
    """Create sample demo data for new users."""
    np.random.seed(42)
    n_schools = 50

    districts = ["Chennai", "Madurai", "Coimbatore", "Trichy", "Salem"]
    blocks = ["Block A", "Block B", "Block C", "Block D"]
    school_types = ["Primary", "Middle", "High School", "Higher Secondary"]
    managements = ["Government", "Aided", "Private"]

    data = {
        "UDISE": [f"33{i:08d}" for i in range(1, n_schools + 1)],
        "School_Name": [f"Demo School {i}" for i in range(1, n_schools + 1)],
        "District": np.random.choice(districts, n_schools),
        "Block": np.random.choice(blocks, n_schools),
        "School_Type": np.random.choice(school_types, n_schools),
        "Management": np.random.choice(managements, n_schools),
    }

    # Add class-wise enrollment data
    for i in range(1, 13):
        data[f"Class{i}_Boys"] = np.random.randint(10, 100, n_schools)
        data[f"Class{i}_Girls"] = np.random.randint(10, 100, n_schools)
        data[f"Class{i}_Transgen"] = np.random.randint(0, 5, n_schools)

    # Add some additional fields
    data["Teachers_Male"] = np.random.randint(5, 30, n_schools)
    data["Teachers_Female"] = np.random.randint(5, 30, n_schools)
    data["Classrooms"] = np.random.randint(10, 40, n_schools)

    return pd.DataFrame(data)

# ═══════════════════════════════════════════════════════════════════════════════
# TRANSLATIONS
# ═══════════════════════════════════════════════════════════════════════════════

TRANSLATIONS = {
    "en": {
        "title": "UDISE Data Generator",
        "subtitle": "Transform, Analyze & Export Educational Data with Ease",
        "welcome_title": "Welcome! Here's what you can do:",
        "welcome_points": [
            "Filter schools by district, block, or management type",
            "Search specific schools using UDISE codes",
            "Create calculated fields (totals, averages, custom formulas)",
            "Pivot data for analysis (sum, count, average by categories)",
            "Export to Excel or CSV with professional formatting"
        ],
        "get_started": "Get Started",
        "step1": "Load Data",
        "step1_desc": "Import data or use demo",
        "step2": "Filter & Select",
        "step2_desc": "Choose schools and apply filters",
        "step3": "Add Fields",
        "step3_desc": "Create calculated columns",
        "step4": "Pivot & Analyze",
        "step4_desc": "Summarize your data",
        "step5": "Export",
        "step5_desc": "Download your results",
        "upload": "Upload Your Data (Optional)",
        "preview": "Data Preview",
        "udise_col": "UDISE Column",
        "udise_input": "Enter UDISE Codes (Optional)",
        "udise_placeholder": "Enter codes separated by comma or new line...\nLeave empty to use all filtered data",
        "select_columns": "Select Output Columns",
        "generate": "Generate Output",
        "filters": "Filters",
        "create_calc": "Create Calculated Fields",
        "calc_type": "Calculation Type",
        "sum": "Sum",
        "diff": "Difference (A - B)",
        "avg": "Average",
        "custom": "Custom Formula",
        "new_field": "New Field Name",
        "add_field": "Add Field",
        "preset_formulas": "Quick Presets",
        "download_excel": "Download Excel",
        "download_csv": "Download CSV",
        "no_file": "No data loaded. Upload a file or use demo data.",
        "no_udise": "No UDISE filter applied - showing all filtered records",
        "no_matches": "No matching UDISE codes found",
        "found_matches": "Found {n} matching records",
        "apply_filters": "Apply Filters",
        "clear_filters": "Clear All",
        "total_records": "Total Records",
        "filtered_records": "Filtered Records",
        "columns": "Columns",
        "selected": "Selected",
        "data_source": "Data Source",
        "online_source": "Online Master",
        "local_source": "Local File",
        "uploaded_source": "Uploaded",
        "demo_source": "Demo Data",
        "help_udise": "Optional: Enter school UDISE codes to extract specific records, or leave empty for all",
        "help_filters": "Use filters to narrow down results",
        "help_presets": "Quick buttons to add common calculated fields",
        "help_custom": "Create your own calculated columns",
        "copy_output": "Copy to Clipboard",
        "workflow_title": "How It Works",
        "class_totals": "Class Totals (1-12)",
        "enrollment_presets": "Enrollment Presets",
        "created_fields": "Created Fields",
        "no_fields": "No custom fields created yet",
        "remove": "Remove",
        "formula": "Formula",
        "select_all": "Select All",
        "deselect_all": "Deselect All",
        "search_columns": "Search columns...",
        "stats": "Statistics",
        "quick_actions": "Quick Actions",
        "pivot_table": "Pivot Table",
        "pivot_desc": "Summarize data by grouping rows and aggregating values",
        "pivot_rows": "Group By (Rows)",
        "pivot_values": "Values to Aggregate",
        "pivot_agg": "Aggregation",
        "pivot_generate": "Generate Pivot",
        "use_demo": "Try Demo Data",
        "use_upload": "Upload Your Data",
        "demo_mode": "Demo Mode",
        "data_mode": "Data Mode"
    },
    "ta": {
        "title": "UDISE தரவு உருவாக்கி",
        "subtitle": "கல்வி தரவை மாற்றுங்கள், பகுப்பாய்வு செய்யுங்கள் & ஏற்றுமதி செய்யுங்கள்",
        "welcome_title": "வரவேற்கிறோம்! நீங்கள் என்ன செய்யலாம்:",
        "welcome_points": [
            "மாவட்டம், தொகுதி அல்லது நிர்வாக வகையால் பள்ளிகளை வடிகட்டுங்கள்",
            "UDISE குறியீடுகளைப் பயன்படுத்தி குறிப்பிட்ட பள்ளிகளைத் தேடுங்கள்",
            "கணக்கிடப்பட்ட புலங்களை உருவாக்குங்கள் (மொத்தங்கள், சராசரிகள், தனிப்பயன் சூத்திரங்கள்)",
            "பகுப்பாய்வுக்காக தரவை பிவோட் செய்யுங்கள்",
            "Excel அல்லது CSV ஆக ஏற்றுமதி செய்யுங்கள்"
        ],
        "get_started": "தொடங்கு",
        "step1": "தரவு ஏற்றுக",
        "step1_desc": "தரவை இறக்குமதி செய்க அல்லது டெமோ பயன்படுத்துக",
        "step2": "வடிகட்டு & தேர்வு",
        "step2_desc": "பள்ளிகளைத் தேர்ந்தெடுத்து வடிகட்டல்களைப் பயன்படுத்துக",
        "step3": "புலங்கள் சேர்",
        "step3_desc": "கணக்கிடப்பட்ட நெடுவரிசைகளை உருவாக்குக",
        "step4": "பிவோட் & பகுப்பாய்வு",
        "step4_desc": "உங்கள் தரவை சுருக்கமாக்குங்கள்",
        "step5": "ஏற்றுமதி",
        "step5_desc": "உங்கள் முடிவுகளைப் பதிவிறக்குக",
        "upload": "உங்கள் தரவை பதிவேற்றுக (விருப்பத்தேர்வு)",
        "preview": "தரவு முன்னோட்டம்",
        "udise_col": "UDISE நெடுவரிசை",
        "udise_input": "UDISE குறியீடுகளை உள்ளிடுக (விருப்பத்தேர்வு)",
        "udise_placeholder": "கமா அல்லது புதிய வரியால் பிரிக்கப்பட்ட குறியீடுகளை உள்ளிடுக...\nஅனைத்து வடிகட்டப்பட்ட தரவையும் பயன்படுத்த காலியாக விடவும்",
        "select_columns": "வெளியீட்டு நெடுவரிசைகளைத் தேர்ந்தெடுக",
        "generate": "வெளியீட்டை உருவாக்கு",
        "filters": "வடிகட்டல்கள்",
        "create_calc": "கணக்கிடப்பட்ட புலங்களை உருவாக்கு",
        "calc_type": "கணக்கீட்டு வகை",
        "sum": "கூட்டல்",
        "diff": "வித்தியாசம் (A - B)",
        "avg": "சராசரி",
        "custom": "தனிப்பயன் சூத்திரம்",
        "new_field": "புதிய புலப் பெயர்",
        "add_field": "புலம் சேர்",
        "preset_formulas": "விரைவு முன்னமைவுகள்",
        "download_excel": "Excel பதிவிறக்கம்",
        "download_csv": "CSV பதிவிறக்கம்",
        "no_file": "தரவு ஏற்றப்படவில்லை. கோப்பை பதிவேற்றுக அல்லது டெமோ தரவை பயன்படுத்துக.",
        "no_udise": "UDISE வடிகட்டல் இல்லை - அனைத்து வடிகட்டப்பட்ட பதிவுகளையும் காட்டுகிறது",
        "no_matches": "பொருந்தக்கூடிய UDISE குறியீடுகள் இல்லை",
        "found_matches": "{n} பொருந்தக்கூடிய பதிவுகள் காணப்பட்டன",
        "apply_filters": "வடிகட்டல்களைப் பயன்படுத்து",
        "clear_filters": "அனைத்தையும் அழி",
        "total_records": "மொத்த பதிவுகள்",
        "filtered_records": "வடிகட்டப்பட்ட பதிவுகள்",
        "columns": "நெடுவரிசைகள்",
        "selected": "தேர்ந்தெடுக்கப்பட்டது",
        "data_source": "தரவு மூலம்",
        "online_source": "ஆன்லைன் மாஸ்டர்",
        "local_source": "உள்ளூர் கோப்பு",
        "uploaded_source": "பதிவேற்றப்பட்டது",
        "demo_source": "டெமோ தரவு",
        "help_udise": "விருப்பத்தேர்வு: குறிப்பிட்ட பதிவுகளுக்கு UDISE குறியீடுகளை உள்ளிடுக, அல்லது அனைத்திற்கும் காலியாக விடுக",
        "help_filters": "முடிவுகளைக் குறைக்க வடிகட்டல்களைப் பயன்படுத்துக",
        "help_presets": "பொதுவான கணக்கிடப்பட்ட புலங்களைச் சேர்க்க விரைவு பொத்தான்கள்",
        "help_custom": "உங்கள் சொந்த கணக்கிடப்பட்ட நெடுவரிசைகளை உருவாக்குக",
        "copy_output": "கிளிப்போர்டுக்கு நகலெடு",
        "workflow_title": "இது எப்படி வேலை செய்கிறது",
        "class_totals": "வகுப்பு மொத்தங்கள் (1-12)",
        "enrollment_presets": "பதிவு முன்னமைவுகள்",
        "created_fields": "உருவாக்கப்பட்ட புலங்கள்",
        "no_fields": "தனிப்பயன் புலங்கள் இன்னும் உருவாக்கப்படவில்லை",
        "remove": "நீக்கு",
        "formula": "சூத்திரம்",
        "select_all": "அனைத்தையும் தேர்",
        "deselect_all": "அனைத்தையும் நீக்கு",
        "search_columns": "நெடுவரிசைகளைத் தேடு...",
        "stats": "புள்ளிவிவரங்கள்",
        "quick_actions": "விரைவு செயல்கள்",
        "pivot_table": "பிவோட் அட்டவணை",
        "pivot_desc": "வரிசைகளை குழுவாக்கி மதிப்புகளை சேர்த்து தரவை சுருக்கமாக்குங்கள்",
        "pivot_rows": "குழுவாக்கு (வரிசைகள்)",
        "pivot_values": "சேர்க்க மதிப்புகள்",
        "pivot_agg": "சேர்க்கை முறை",
        "pivot_generate": "பிவோட் உருவாக்கு",
        "use_demo": "டெமோ தரவை முயற்சிக்கவும்",
        "use_upload": "உங்கள் தரவை பதிவேற்றுக",
        "demo_mode": "டெமோ முறை",
        "data_mode": "தரவு முறை"
    }
}

# ═══════════════════════════════════════════════════════════════════════════════
# SESSION STATE INITIALIZATION
# ═══════════════════════════════════════════════════════════════════════════════

if "formula_presets" not in st.session_state:
    st.session_state["formula_presets"] = {}
if "extra_fields" not in st.session_state:
    st.session_state["extra_fields"] = []
if "created_fields" not in st.session_state:
    st.session_state["created_fields"] = {}
if "selected_columns" not in st.session_state:
    st.session_state["selected_columns"] = []
if "data_loaded" not in st.session_state:
    st.session_state["data_loaded"] = False
if "show_welcome" not in st.session_state:
    st.session_state["show_welcome"] = True
if "use_demo" not in st.session_state:
    st.session_state["use_demo"] = False
if "pivot_result" not in st.session_state:
    st.session_state["pivot_result"] = None

# ═══════════════════════════════════════════════════════════════════════════════
# SIDEBAR - LANGUAGE & SETTINGS
# ═══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("### ⚙️ Settings")
    lang_choice = st.radio(
        "Language / மொழி",
        ("English", "தமிழ்"),
        horizontal=True,
        label_visibility="collapsed"
    )
    lang = "en" if lang_choice == "English" else "ta"
    tr = TRANSLATIONS[lang]

    st.markdown("---")

# ═══════════════════════════════════════════════════════════════════════════════
# HEADER
# ═══════════════════════════════════════════════════════════════════════════════

st.markdown(f"""
<div class="main-header">
    <h1>🏫 {tr['title']}</h1>
    <p>{tr['subtitle']}</p>
</div>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# WELCOME SECTION - What this tool does
# ═══════════════════════════════════════════════════════════════════════════════

if st.session_state["show_welcome"]:
    st.markdown(f"""
    <div class="welcome-card">
        <h3>👋 {tr['welcome_title']}</h3>
        <ul>
            {"".join([f"<li>{point}</li>" for point in tr['welcome_points']])}
        </ul>
    </div>
    """, unsafe_allow_html=True)

    # Feature cards
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-icon">🔍</div>
            <h4>Filter & Search</h4>
            <p>Filter by district, block, or search by UDISE codes</p>
        </div>
        """, unsafe_allow_html=True)
    with col2:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-icon">📊</div>
            <h4>Create Fields</h4>
            <p>Add calculated columns like totals and averages</p>
        </div>
        """, unsafe_allow_html=True)
    with col3:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-icon">📈</div>
            <h4>Pivot Analysis</h4>
            <p>Summarize data by grouping and aggregating</p>
        </div>
        """, unsafe_allow_html=True)
    with col4:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-icon">📥</div>
            <h4>Export</h4>
            <p>Download as formatted Excel or CSV</p>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("")

    # Hide welcome button
    if st.button("✨ Got it! Let's start", type="primary", use_container_width=True):
        st.session_state["show_welcome"] = False
        st.experimental_rerun()

    st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# DATA SOURCE SELECTION
# ═══════════════════════════════════════════════════════════════════════════════

MASTER_URL = "https://d3ijhv7dn0xr3b.cloudfront.net/10684.csv"

df_master = None
source_used = None
source_type = None

# Data source selection
st.markdown("### 📂 Data Source")
data_col1, data_col2 = st.columns(2)

with data_col1:
    if st.button(f"🎮 {tr['use_demo']}", use_container_width=True,
                 type="primary" if st.session_state["use_demo"] else "secondary",
                 help="Load sample data to explore the tool"):
        st.session_state["use_demo"] = True
        st.experimental_rerun()

with data_col2:
    if st.button(f"📤 {tr['use_upload']}", use_container_width=True,
                 type="primary" if not st.session_state["use_demo"] else "secondary",
                 help="Upload your own Excel/CSV file"):
        st.session_state["use_demo"] = False
        st.experimental_rerun()

st.markdown("")

# Load data based on selection
if st.session_state["use_demo"]:
    df_master = create_demo_data()
    source_used = "Demo Data (50 sample schools)"
    source_type = "demo"
    st.markdown(f"""
    <div class="info-box">
        🎮 <strong>Demo Mode Active</strong> - Using sample data with 50 schools.
        Perfect for exploring features! Switch to "Upload Your Data" to use real data.
    </div>
    """, unsafe_allow_html=True)
else:
    # Try loading from online master URL first
    try:
        with st.spinner("🔄 Fetching master data from online source..."):
            response = requests.get(MASTER_URL, timeout=10)
            if response.status_code == 200:
                data = BytesIO(response.content)
                if MASTER_URL.lower().endswith(".csv"):
                    df_master = pd.read_csv(data, dtype=str)
                elif MASTER_URL.lower().endswith(".xls"):
                    df_master = pd.read_excel(data, engine="xlrd", dtype=str)
                else:
                    df_master = pd.read_excel(data, engine="openpyxl", dtype=str)
                source_used = "Online Master Database"
                source_type = "online"
    except Exception as e:
        pass

    # Try local default master files (fallback)
    if df_master is None:
        default_files = ["master.xlsx", "master.xls", "master.csv"]
        for f in default_files:
            if os.path.exists(f):
                try:
                    if f.endswith(".csv"):
                        df_master = pd.read_csv(f, dtype=str)
                    elif f.endswith(".xls"):
                        df_master = pd.read_excel(f, engine="xlrd", dtype=str)
                    else:
                        df_master = pd.read_excel(f, engine="openpyxl", dtype=str)
                    source_used = f"Local: {f}"
                    source_type = "local"
                    break
                except Exception as e:
                    pass

    # File upload option (always available)
    with st.expander(f"📁 {tr['upload']}", expanded=(df_master is None)):
        uploaded_file = st.file_uploader(
            "Drop your file here or click to browse",
            type=["xlsx", "xls", "csv"],
            label_visibility="collapsed"
        )

        if uploaded_file is not None:
            try:
                fname = uploaded_file.name.lower()
                if fname.endswith(".csv"):
                    df_master = pd.read_csv(uploaded_file, dtype=str)
                elif fname.endswith(".xls"):
                    df_master = pd.read_excel(uploaded_file, engine="xlrd", dtype=str)
                else:
                    df_master = pd.read_excel(uploaded_file, engine="openpyxl", dtype=str)
                source_used = f"Uploaded: {uploaded_file.name}"
                source_type = "uploaded"
                st.success(f"✅ File loaded successfully!")
            except Exception as e:
                st.error(f"❌ Error reading file: {e}")

# Final check - if no data, offer demo
if df_master is None:
    st.markdown(f"""
    <div class="warning-box">
        ⚠️ <strong>No data loaded.</strong> Click "Try Demo Data" above to explore the tool with sample data,
        or upload your own Excel/CSV file.
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# Normalize columns
df_master.columns = df_master.columns.str.strip()

# Coerce class gender columns to numeric
for col in df_master.columns:
    if re.match(r"(?i)^Class\d+_(Boys|Girls|Transgen)$", col):
        df_master[col] = pd.to_numeric(df_master[col], errors="coerce").fillna(0)

# Working copy
df = df_master.copy()

# ═══════════════════════════════════════════════════════════════════════════════
# DATA STATISTICS DISPLAY
# ═══════════════════════════════════════════════════════════════════════════════

# Display source and stats
source_icon = {"online": "🌐", "local": "💾", "uploaded": "📤", "demo": "🎮"}.get(source_type, "📊")
st.markdown(f"""
<div class="success-box">
    <strong>{source_icon} {tr['data_source']}:</strong> {source_used}
    {f'<span class="demo-badge">DEMO</span>' if source_type == "demo" else ""}
</div>
""", unsafe_allow_html=True)

stat_col1, stat_col2, stat_col3, stat_col4 = st.columns(4)
with stat_col1:
    st.markdown(f"""
    <div class="stat-card">
        <div class="stat-value">{len(df_master):,}</div>
        <div class="stat-label">{tr['total_records']}</div>
    </div>
    """, unsafe_allow_html=True)
with stat_col2:
    st.markdown(f"""
    <div class="stat-card">
        <div class="stat-value">{len(df.columns)}</div>
        <div class="stat-label">{tr['columns']}</div>
    </div>
    """, unsafe_allow_html=True)
with stat_col3:
    st.markdown(f"""
    <div class="stat-card">
        <div class="stat-value">{len(st.session_state.get('extra_fields', []))}</div>
        <div class="stat-label">{tr['created_fields']}</div>
    </div>
    """, unsafe_allow_html=True)
with stat_col4:
    st.markdown(f"""
    <div class="stat-card">
        <div class="stat-value">{len(st.session_state['selected_columns'])}</div>
        <div class="stat-label">{tr['selected']}</div>
    </div>
    """, unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# SIDEBAR FILTERS - DYNAMIC FOR ALL COLUMNS
# ═══════════════════════════════════════════════════════════════════════════════

def get_filterable_columns(dataframe, max_unique=100):
    """Get columns suitable for filtering (categorical or low-cardinality)."""
    filterable = []
    for col in dataframe.columns:
        unique_count = dataframe[col].nunique()
        # Include columns with reasonable number of unique values for filtering
        if unique_count <= max_unique and unique_count > 1:
            filterable.append((col, unique_count))
    # Sort by number of unique values (ascending)
    filterable.sort(key=lambda x: x[1])
    return [col for col, _ in filterable]

selected_filters = {}
with st.sidebar:
    st.markdown(f"### 🔍 {tr['filters']}")
    st.caption(tr['help_filters'])

    # Get all filterable columns dynamically
    filterable_columns = get_filterable_columns(df)

    if filterable_columns:
        # Column selector for filters
        st.markdown("**Select columns to filter:**")
        selected_filter_cols = st.multiselect(
            "Choose filter columns",
            options=filterable_columns,
            default=[],
            key="dynamic_filter_columns",
            label_visibility="collapsed",
            help="Select which columns you want to filter by"
        )

        st.markdown("---")

        if selected_filter_cols:
            with st.form("filters_form"):
                for col in selected_filter_cols:
                    options = sorted(df[col].dropna().astype(str).unique().tolist())
                    chosen = st.multiselect(
                        f"📌 {col}",
                        options=options,
                        key=f"filter_{col}",
                        help=f"Filter by {col} ({len(options)} values)"
                    )
                    if chosen:
                        selected_filters[col] = chosen

                filter_cols = st.columns(2)
                with filter_cols[0]:
                    apply_filters = st.form_submit_button(f"✅ {tr['apply_filters']}", use_container_width=True)
                with filter_cols[1]:
                    clear_btn = st.form_submit_button(f"🗑️ {tr['clear_filters']}", use_container_width=True)
        else:
            st.info("👆 Select columns above to enable filtering")
    else:
        st.info("No filterable columns detected in data")

# Apply filters
if selected_filters:
    mask = pd.Series([True] * len(df))
    for col, vals in selected_filters.items():
        mask = mask & df[col].astype(str).isin(vals)
    df = df[mask]

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN CONTENT TABS
# ═══════════════════════════════════════════════════════════════════════════════

st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    f"🔍 {tr['step2']}",
    f"➕ {tr['step3']}",
    f"📈 {tr['pivot_table']}",
    f"🔄 Compare & Match",
    f"📋 {tr['select_columns']}",
    f"👀 {tr['preview']}"
])

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1: UDISE SELECTION (OPTIONAL)
# ═══════════════════════════════════════════════════════════════════════════════

with tab1:
    st.markdown(f"### 🏫 {tr['udise_input']}")
    st.caption(tr['help_udise'])

    # UDISE column auto-detect
    udise_candidates = ["UDISE", "UDISE Code", "UDISE_Code", "udise", "udise_code", "UDISECODE"]
    udise_col = find_column(df, udise_candidates)

    col1, col2 = st.columns([3, 1])
    with col1:
        if not udise_col:
            udise_col_options = ["(None - Skip UDISE filtering)"] + list(df.columns)
            udise_col_selected = st.selectbox(
                tr['udise_col'],
                options=udise_col_options,
                help="Select the column containing UDISE codes, or skip to use all data"
            )
            udise_col = None if udise_col_selected == "(None - Skip UDISE filtering)" else udise_col_selected
        else:
            st.info(f"📌 Auto-detected UDISE column: **{udise_col}**")

    with col2:
        st.metric("Available Schools", f"{len(df):,}")

    udise_input = st.text_area(
        tr['udise_input'],
        height=120,
        placeholder=tr['udise_placeholder'],
        label_visibility="collapsed"
    )

    udise_list = []
    if udise_input and udise_col:
        udise_list = [u.strip() for u in udise_input.replace("\r", "\n").replace(",", "\n").split("\n") if u.strip()]

    # Apply UDISE filter only if column is selected AND codes are provided
    if udise_list and udise_col:
        df = df[df[udise_col].astype(str).isin(udise_list)]

        # Maintain user-given UDISE order
        try:
            df[udise_col] = df[udise_col].astype(str)
            df = df.set_index(udise_col).loc[[u for u in udise_list if u in df.index]].reset_index()
        except Exception:
            pass

        if len(df) > 0:
            st.markdown(f"""
            <div class="success-box">
                ✅ <strong>{tr['found_matches'].format(n=len(df))}</strong> from {len(udise_list)} codes entered
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div class="warning-box">
                ⚠️ <strong>{tr['no_matches']}</strong>
            </div>
            """, unsafe_allow_html=True)
    else:
        if not udise_col:
            st.markdown(f"""
            <div class="info-box">
                💡 <strong>UDISE filtering skipped</strong> - showing all filtered records ({len(df):,} schools available)
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div class="info-box">
                💡 <strong>{tr['no_udise']}</strong> ({len(df):,} schools available)
            </div>
            """, unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2: CALCULATED FIELDS
# ═══════════════════════════════════════════════════════════════════════════════

with tab2:
    st.markdown(f"### ⚡ {tr['quick_actions']}")
    st.caption(tr['help_presets'])

    # Quick preset buttons
    preset_col1, preset_col2, preset_col3 = st.columns(3)

    with preset_col1:
        if st.button(f"📊 {tr['class_totals']}", use_container_width=True, help="Creates Class1_Total through Class12_Total"):
            created = build_class_totals(df)
            for cname in created:
                if cname not in st.session_state.get("extra_fields", []):
                    st.session_state.setdefault("extra_fields", []).append(cname)
            st.success(f"✅ Created {len(created)} class total fields!")
            st.experimental_rerun()

    with preset_col2:
        if st.button(f"📈 {tr['enrollment_presets']}", use_container_width=True, help="Creates enrollment aggregations by grade ranges"):
            build_class_totals(df)
            created = build_enrollment_presets(df)
            for cname in created:
                if cname not in st.session_state.get("extra_fields", []):
                    st.session_state.setdefault("extra_fields", []).append(cname)
            st.success(f"✅ Created {len(created)} enrollment fields!")
            st.experimental_rerun()

    with preset_col3:
        if st.button("🔄 Clear All Fields", use_container_width=True, type="secondary"):
            st.session_state["extra_fields"] = []
            st.session_state["created_fields"] = {}
            st.experimental_rerun()

    st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

    # Show created fields
    if st.session_state.get("extra_fields", []):
        st.markdown(f"### 📦 {tr['created_fields']}")

        # Display as chips
        field_html = ""
        for field in st.session_state.get("extra_fields", []):
            field_html += f'<span class="field-chip">{field}</span>'
        st.markdown(f'<div>{field_html}</div>', unsafe_allow_html=True)
        st.markdown("")

    st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

    # Custom calculated fields
    st.markdown(f"### 🔧 {tr['create_calc']}")
    st.caption(tr['help_custom'])

    numeric_candidates = get_numeric_columns(df)

    # Add existing calculated fields to numeric candidates
    for f in st.session_state.get("extra_fields", []):
        if f not in numeric_candidates:
            numeric_candidates.append(f)

    calc_col1, calc_col2 = st.columns([1, 1])

    with calc_col1:
        calc_type = st.selectbox(
            tr['calc_type'],
            [tr["sum"], tr["diff"], tr["avg"], tr["custom"]],
            help="Choose how to calculate the new field"
        )

    with calc_col2:
        new_field_name = st.text_input(
            tr['new_field'],
            placeholder="e.g., Total_Students",
            help="Name for your new calculated field"
        )

    # Dynamic inputs based on calculation type
    if calc_type == tr["diff"]:
        diff_col1, diff_col2 = st.columns(2)
        with diff_col1:
            col_a = st.selectbox("Column A", options=numeric_candidates, key="diffA")
        with diff_col2:
            col_b = st.selectbox("Column B (subtract)", options=numeric_candidates, key="diffB")
        st.caption(f"Formula: {col_a} - {col_b}")

    elif calc_type in (tr["sum"], tr["avg"]):
        cols_to_use = st.multiselect(
            "Select columns to " + ("sum" if calc_type == tr["sum"] else "average"),
            options=numeric_candidates,
            key="sum_cols",
            help="Select one or more numeric columns"
        )
        if cols_to_use:
            st.caption(f"Formula: {' + '.join(cols_to_use)}" + (f" / {len(cols_to_use)}" if calc_type == tr["avg"] else ""))

    else:  # Custom formula
        st.caption("Use column names and operators (+, -, *, /, parentheses)")
        custom_formula = st.text_input(
            tr['formula'],
            placeholder="(Class1_Total + Class2_Total) / Total_Enrollment",
            key="custom_formula"
        )

    if st.button(f"➕ {tr['add_field']}", type="primary", use_container_width=True):
        if not new_field_name:
            st.error("❌ Please enter a name for the new field")
        elif new_field_name in df.columns or new_field_name in st.session_state.get("extra_fields", []):
            st.error(f"❌ Field '{new_field_name}' already exists")
        else:
            try:
                if calc_type == tr["diff"]:
                    a = pd.to_numeric(df[col_a], errors="coerce").fillna(0)
                    b = pd.to_numeric(df[col_b], errors="coerce").fillna(0)
                    df[new_field_name] = a - b
                    meta = ("diff", (col_a, col_b))
                elif calc_type == tr["sum"]:
                    if not cols_to_use:
                        st.error("❌ Select at least one column")
                        raise RuntimeError("no cols")
                    df[new_field_name] = safe_numeric_sum(df, cols_to_use)
                    meta = ("sum", cols_to_use)
                elif calc_type == tr["avg"]:
                    if not cols_to_use:
                        st.error("❌ Select at least one column")
                        raise RuntimeError("no cols")
                    df[new_field_name] = safe_numeric_sum(df, cols_to_use) / len(cols_to_use)
                    meta = ("avg", cols_to_use)
                else:
                    expr = custom_formula.strip()
                    if not expr:
                        st.error("❌ Enter a formula")
                        raise RuntimeError("no formula")
                    env = {c: pd.to_numeric(df[c], errors="coerce").fillna(0) for c in df.columns}
                    df[new_field_name] = eval(expr, {"__builtins__": {}}, env)
                    meta = ("custom", expr)

                st.session_state.setdefault("extra_fields", []).append(new_field_name)
                st.session_state["created_fields"][new_field_name] = {"type": meta[0], "definition": meta[1]}
                st.success(f"✅ Field '{new_field_name}' created successfully!")
                st.experimental_rerun()
            except Exception as e:
                st.error(f"❌ Error: {e}")

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 3: PIVOT TABLE - ENHANCED WITH INDIVIDUAL AGGREGATIONS
# ═══════════════════════════════════════════════════════════════════════════════

with tab3:
    st.markdown(f"### 📈 {tr['pivot_table']}")
    st.caption(tr['pivot_desc'])

    st.markdown(f"""
    <div class="info-box">
        💡 <strong>What is a Pivot Table?</strong><br>
        A pivot table lets you summarize large amounts of data. For example:
        "Show total enrollment grouped by District" or "Count schools by Management type".<br>
        <strong>NEW:</strong> You can now select different aggregation types for each value column!
    </div>
    """, unsafe_allow_html=True)

    # Get categorical and numeric columns
    categorical_cols = []
    numeric_cols = []
    all_cols_for_count = list(df.columns)  # All columns available for distinct count

    for col in df.columns:
        if df[col].dtype == 'object' or df[col].nunique() < 50:
            categorical_cols.append(col)
        tmp = pd.to_numeric(df[col], errors='coerce')
        if not tmp.isnull().all():
            numeric_cols.append(col)

    # Add created fields to numeric cols
    for f in st.session_state.get("extra_fields", []):
        if f not in numeric_cols:
            numeric_cols.append(f)

    # All aggregation options including distinct count
    AGG_OPTIONS = ["Sum", "Count", "Distinct Count", "Average", "Min", "Max", "Median", "Std Dev", "First", "Last"]

    pivot_col1, pivot_col2 = st.columns(2)

    with pivot_col1:
        pivot_rows = st.multiselect(
            tr['pivot_rows'],
            options=categorical_cols,
            help="Select columns to group your data by (e.g., District, Block)",
            max_selections=5
        )

    with pivot_col2:
        pivot_values = st.multiselect(
            tr['pivot_values'],
            options=numeric_cols + [c for c in all_cols_for_count if c not in numeric_cols],
            help="Select columns to aggregate (numeric for most, any for Count/Distinct Count)"
        )

    # Individual aggregation selection for each value column
    st.markdown("#### 🎯 Select Aggregation for Each Value Column")
    st.caption("Choose how to aggregate each selected value column")

    pivot_agg_map = {}
    if pivot_values:
        agg_cols = st.columns(min(len(pivot_values), 4))
        for idx, val_col in enumerate(pivot_values):
            col_idx = idx % 4
            with agg_cols[col_idx]:
                selected_agg = st.selectbox(
                    f"{val_col[:20]}{'...' if len(val_col) > 20 else ''}",
                    options=AGG_OPTIONS,
                    key=f"pivot_agg_{val_col}",
                    help=f"Aggregation for {val_col}"
                )
                pivot_agg_map[val_col] = selected_agg
    else:
        st.info("👆 Select value columns above to configure aggregations")

    st.markdown("---")

    if st.button(f"📊 {tr['pivot_generate']}", type="primary", use_container_width=True):
        if not pivot_rows:
            st.error("❌ Please select at least one column to group by")
        elif not pivot_values:
            st.error("❌ Please select at least one value column to aggregate")
        else:
            try:
                df_pivot = df.copy()

                # Prepare aggregation functions
                agg_funcs = {}
                for val_col, agg_type in pivot_agg_map.items():
                    # Convert to numeric for most aggregations (except Count/Distinct Count on text)
                    if agg_type not in ["Count", "Distinct Count", "First", "Last"]:
                        df_pivot[val_col] = pd.to_numeric(df_pivot[val_col], errors='coerce').fillna(0)

                    if agg_type == "Sum":
                        agg_funcs[val_col] = 'sum'
                    elif agg_type == "Count":
                        agg_funcs[val_col] = 'count'
                    elif agg_type == "Distinct Count":
                        agg_funcs[val_col] = 'nunique'
                    elif agg_type == "Average":
                        agg_funcs[val_col] = 'mean'
                    elif agg_type == "Min":
                        agg_funcs[val_col] = 'min'
                    elif agg_type == "Max":
                        agg_funcs[val_col] = 'max'
                    elif agg_type == "Median":
                        agg_funcs[val_col] = 'median'
                    elif agg_type == "Std Dev":
                        agg_funcs[val_col] = 'std'
                    elif agg_type == "First":
                        agg_funcs[val_col] = 'first'
                    elif agg_type == "Last":
                        agg_funcs[val_col] = 'last'

                # Create pivot table with individual aggregations
                pivot_result = df_pivot.groupby(pivot_rows).agg(agg_funcs).reset_index()

                # Rename columns to include aggregation type
                new_columns = {}
                for col in pivot_result.columns:
                    if col in pivot_agg_map:
                        new_columns[col] = f"{col}_{pivot_agg_map[col].replace(' ', '_')}"
                pivot_result = pivot_result.rename(columns=new_columns)

                # Round numeric columns
                for col in pivot_result.columns:
                    if pivot_result[col].dtype in ['float64', 'float32']:
                        pivot_result[col] = pivot_result[col].round(2)

                st.session_state["pivot_result"] = pivot_result
                st.success(f"✅ Pivot table created with {len(pivot_result)} rows!")

            except Exception as e:
                st.error(f"❌ Error creating pivot: {e}")

    # Display pivot result
    if st.session_state["pivot_result"] is not None:
        st.markdown("### 📋 Pivot Result")
        st.dataframe(st.session_state["pivot_result"], use_container_width=True, height=400)

        # Download pivot
        pivot_excel = to_excel_bytes_styled(st.session_state["pivot_result"])
        pivot_csv = st.session_state["pivot_result"].to_csv(index=False).encode("utf-8")

        dl_col1, dl_col2 = st.columns(2)
        with dl_col1:
            st.download_button(
                f"📗 Download Pivot (Excel)",
                data=pivot_excel,
                file_name="UDISE_Pivot.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with dl_col2:
            st.download_button(
                f"📄 Download Pivot (CSV)",
                data=pivot_csv,
                file_name="UDISE_Pivot.csv",
                mime="text/csv",
                use_container_width=True
            )

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 4: COMPARE & MATCH FILES
# ═══════════════════════════════════════════════════════════════════════════════

# Initialize session state for comparison
if "comparison_file" not in st.session_state:
    st.session_state["comparison_file"] = None
if "comparison_result" not in st.session_state:
    st.session_state["comparison_result"] = None

with tab4:
    st.markdown("### 🔄 Compare & Match Data")
    st.caption("Upload a comparison file to find matched and unmatched records")

    st.markdown(f"""
    <div class="info-box">
        💡 <strong>How it works:</strong><br>
        1. Upload a comparison file (Excel/CSV) containing values to match<br>
        2. Select the matching column from both master data and comparison file<br>
        3. Generate lists of matched and not-matched records
    </div>
    """, unsafe_allow_html=True)

    # Comparison file upload
    st.markdown("#### 📁 Upload Comparison File (Optional)")
    compare_file = st.file_uploader(
        "Upload comparison file",
        type=["xlsx", "xls", "csv"],
        key="compare_file_uploader",
        label_visibility="collapsed",
        help="Upload an Excel or CSV file to compare against master data"
    )

    df_compare = None
    if compare_file is not None:
        try:
            fname = compare_file.name.lower()
            if fname.endswith(".csv"):
                df_compare = pd.read_csv(compare_file, dtype=str)
            elif fname.endswith(".xls"):
                df_compare = pd.read_excel(compare_file, engine="xlrd", dtype=str)
            else:
                df_compare = pd.read_excel(compare_file, engine="openpyxl", dtype=str)

            df_compare.columns = df_compare.columns.str.strip()
            st.session_state["comparison_file"] = df_compare

            st.markdown(f"""
            <div class="success-box">
                ✅ <strong>Comparison file loaded:</strong> {compare_file.name}<br>
                📊 {len(df_compare)} records, {len(df_compare.columns)} columns
            </div>
            """, unsafe_allow_html=True)
        except Exception as e:
            st.error(f"❌ Error reading comparison file: {e}")

    # Use stored comparison file if available
    if df_compare is None and st.session_state["comparison_file"] is not None:
        df_compare = st.session_state["comparison_file"]

    if df_compare is not None:
        st.markdown("---")
        st.markdown("#### 🎯 Configure Matching")

        match_col1, match_col2 = st.columns(2)

        with match_col1:
            st.markdown("**Master Data Column:**")
            master_match_col = st.selectbox(
                "Select column from master data",
                options=list(df.columns),
                key="master_match_column",
                label_visibility="collapsed",
                help="Column in your master/filtered data to match against"
            )

        with match_col2:
            st.markdown("**Comparison File Column:**")
            compare_match_col = st.selectbox(
                "Select column from comparison file",
                options=list(df_compare.columns),
                key="compare_match_column",
                label_visibility="collapsed",
                help="Column in the uploaded comparison file"
            )

        # Output options
        st.markdown("#### 📤 Output Options")
        output_options = st.multiselect(
            "Select which outputs to generate:",
            options=["Matched Records", "Not Matched (in Master)", "Not Matched (in Comparison)", "Full Comparison Report"],
            default=["Matched Records", "Not Matched (in Master)"],
            help="Choose what outputs you want to generate"
        )

        # Include additional columns from comparison file
        st.markdown("**Include columns from comparison file (optional):**")
        extra_compare_cols = st.multiselect(
            "Additional columns from comparison file",
            options=[c for c in df_compare.columns if c != compare_match_col],
            key="extra_compare_columns",
            label_visibility="collapsed",
            help="Select additional columns from comparison file to include in matched output"
        )

        if st.button("🔍 Generate Comparison", type="primary", use_container_width=True):
            try:
                # Clean and prepare data for matching
                master_values = df[master_match_col].astype(str).str.strip()
                compare_values = df_compare[compare_match_col].astype(str).str.strip()

                master_set = set(master_values)
                compare_set = set(compare_values)

                # Calculate matches
                matched_values = master_set & compare_set
                master_only = master_set - compare_set
                compare_only = compare_set - master_set

                results = {}

                # Matched Records
                if "Matched Records" in output_options:
                    matched_df = df[master_values.isin(matched_values)].copy()

                    # Add extra columns from comparison file if requested
                    if extra_compare_cols:
                        compare_subset = df_compare[[compare_match_col] + extra_compare_cols].copy()
                        compare_subset[compare_match_col] = compare_subset[compare_match_col].astype(str).str.strip()
                        matched_df = matched_df.merge(
                            compare_subset,
                            left_on=master_match_col,
                            right_on=compare_match_col,
                            how='left',
                            suffixes=('', '_compare')
                        )

                    matched_df['_match_status'] = 'Matched'
                    results["matched"] = matched_df

                # Not Matched in Master (records in master not found in comparison)
                if "Not Matched (in Master)" in output_options:
                    not_in_compare = df[master_values.isin(master_only)].copy()
                    not_in_compare['_match_status'] = 'Not in Comparison File'
                    results["not_matched_master"] = not_in_compare

                # Not Matched in Comparison (records in comparison not found in master)
                if "Not Matched (in Comparison)" in output_options:
                    not_in_master = df_compare[compare_values.isin(compare_only)].copy()
                    not_in_master['_match_status'] = 'Not in Master Data'
                    results["not_matched_compare"] = not_in_master

                # Full Comparison Report
                if "Full Comparison Report" in output_options:
                    report_data = {
                        'Metric': [
                            'Total Master Records',
                            'Total Comparison Records',
                            'Matched Records',
                            'Not Matched (Master Only)',
                            'Not Matched (Comparison Only)',
                            'Match Rate (Master)',
                            'Match Rate (Comparison)'
                        ],
                        'Value': [
                            len(master_set),
                            len(compare_set),
                            len(matched_values),
                            len(master_only),
                            len(compare_only),
                            f"{(len(matched_values)/len(master_set)*100):.1f}%" if master_set else "N/A",
                            f"{(len(matched_values)/len(compare_set)*100):.1f}%" if compare_set else "N/A"
                        ]
                    }
                    results["report"] = pd.DataFrame(report_data)

                st.session_state["comparison_result"] = results

                # Summary
                st.markdown(f"""
                <div class="success-box">
                    ✅ <strong>Comparison Complete!</strong><br>
                    🔗 Matched: <strong>{len(matched_values)}</strong> records<br>
                    📌 Master Only: <strong>{len(master_only)}</strong> records<br>
                    📌 Comparison Only: <strong>{len(compare_only)}</strong> records
                </div>
                """, unsafe_allow_html=True)

            except Exception as e:
                st.error(f"❌ Error during comparison: {e}")

        # Display and download results
        if st.session_state["comparison_result"]:
            results = st.session_state["comparison_result"]

            st.markdown("---")
            st.markdown("### 📋 Comparison Results")

            result_tabs = st.tabs([k.replace("_", " ").title() for k in results.keys()])

            for idx, (key, result_df) in enumerate(results.items()):
                with result_tabs[idx]:
                    st.dataframe(result_df, use_container_width=True, height=300)
                    st.caption(f"{len(result_df)} records")

                    # Download buttons
                    dl_col1, dl_col2 = st.columns(2)
                    with dl_col1:
                        excel_data = to_excel_bytes_styled(result_df)
                        st.download_button(
                            f"📗 Download {key.replace('_', ' ').title()} (Excel)",
                            data=excel_data,
                            file_name=f"comparison_{key}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key=f"dl_excel_{key}"
                        )
                    with dl_col2:
                        csv_data = result_df.to_csv(index=False).encode("utf-8")
                        st.download_button(
                            f"📄 Download {key.replace('_', ' ').title()} (CSV)",
                            data=csv_data,
                            file_name=f"comparison_{key}.csv",
                            mime="text/csv",
                            use_container_width=True,
                            key=f"dl_csv_{key}"
                        )

    else:
        st.markdown("""
        <div class="warning-box">
            👆 <strong>Upload a comparison file above to get started</strong><br>
            You can upload an Excel (.xlsx, .xls) or CSV file containing values to compare against your master data.
        </div>
        """, unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 5: COLUMN SELECTION
# ═══════════════════════════════════════════════════════════════════════════════

with tab5:
    st.markdown(f"### 📋 {tr['select_columns']}")
    st.caption("Choose which columns to include in your final output")

    # Build available columns
    all_columns = list(df.columns) + [f for f in st.session_state.get("extra_fields", []) if f not in df.columns]
    seen = set()
    available_columns = []
    for c in all_columns:
        if c not in seen:
            available_columns.append(c)
            seen.add(c)

    # Quick actions
    action_col1, action_col2, action_col3 = st.columns([1, 1, 2])
    with action_col1:
        if st.button(f"✅ {tr['select_all']}", use_container_width=True):
            st.session_state["selected_columns"] = available_columns.copy()
            st.experimental_rerun()
    with action_col2:
        if st.button(f"❌ {tr['deselect_all']}", use_container_width=True):
            st.session_state["selected_columns"] = []
            st.experimental_rerun()
    with action_col3:
        search_term = st.text_input(tr['search_columns'], label_visibility="collapsed", placeholder=f"🔍 {tr['search_columns']}")

    # Filter columns by search
    display_columns = available_columns
    if search_term:
        display_columns = [c for c in available_columns if search_term.lower() in c.lower()]

    # Keep previous selections if still available
    default_sel = [c for c in st.session_state["selected_columns"] if c in available_columns]

    selected_columns = st.multiselect(
        tr['select_columns'],
        options=display_columns,
        default=[c for c in default_sel if c in display_columns],
        key="ui_selected_columns",
        label_visibility="collapsed",
        help="Select columns to include in your output"
    )

    # Update session state with new selections while preserving others not in current display
    other_selected = [c for c in st.session_state["selected_columns"] if c not in display_columns]
    st.session_state["selected_columns"] = selected_columns + other_selected

    # Show selection summary
    st.markdown(f"""
    <div class="info-box">
        📊 <strong>{len(st.session_state['selected_columns'])}</strong> columns selected out of <strong>{len(available_columns)}</strong> available
    </div>
    """, unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 6: DATA PREVIEW
# ═══════════════════════════════════════════════════════════════════════════════

with tab6:
    st.markdown(f"### 👀 {tr['preview']}")

    if len(df) > 0:
        st.dataframe(
            df.head(100),
            use_container_width=True,
            height=400
        )
        st.caption(f"Showing first 100 of {len(df):,} records")
    else:
        st.info("No data to preview. Apply filters and enter UDISE codes to see data.")

# ═══════════════════════════════════════════════════════════════════════════════
# GENERATE & EXPORT SECTION
# ═══════════════════════════════════════════════════════════════════════════════

st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
st.markdown("### 🚀 Generate & Export")

generate_col1, generate_col2 = st.columns([3, 1])

with generate_col1:
    generate_btn = st.button(
        f"⚡ {tr['generate']}",
        type="primary",
        use_container_width=True
    )

with generate_col2:
    st.metric("Selected Columns", len(st.session_state['selected_columns']))

if generate_btn:
    if df.empty:
        st.error(f"❌ {tr['no_matches']}")
    elif not st.session_state["selected_columns"]:
        st.error("❌ Please select at least one column for output (go to 'Select Output Columns' tab)")
    else:
        with st.spinner("Generating output..."):
            # Recreate preset/class totals & user-created calculated fields
            class_total_names = [f"Class{i}_Total" for i in range(1, 13)]
            if any(name in st.session_state.get("extra_fields", []) for name in class_total_names):
                build_class_totals(df)

            enroll_preset_names = ["Enrollment_1_5", "Enrollment_6_8", "Enrollment_9_10", "Enrollment_11_12", "Total_Enrollment"]
            if any(name in st.session_state.get("extra_fields", []) for name in enroll_preset_names):
                build_enrollment_presets(df)

            # Recreate user-created fields
            for fname, meta in st.session_state["created_fields"].items():
                if meta["type"] == "diff":
                    a, b = meta["definition"]
                    df[fname] = pd.to_numeric(df.get(a, pd.Series(0, index=df.index)), errors="coerce").fillna(0) - \
                               pd.to_numeric(df.get(b, pd.Series(0, index=df.index)), errors="coerce").fillna(0)
                elif meta["type"] == "sum":
                    df[fname] = safe_numeric_sum(df, meta["definition"])
                elif meta["type"] == "avg":
                    df[fname] = safe_numeric_sum(df, meta["definition"]) / max(1, len(meta["definition"]))
                elif meta["type"] == "custom":
                    env = {c: pd.to_numeric(df.get(c, 0), errors="coerce").fillna(0) for c in df.columns}
                    try:
                        df[fname] = eval(meta["definition"], {"__builtins__": {}}, env)
                    except Exception:
                        df[fname] = pd.Series(0, index=df.index)

            # Validate and get output
            valid_selected = [c for c in st.session_state["selected_columns"] if c in df.columns]
            missing = [c for c in st.session_state["selected_columns"] if c not in df.columns]

            if missing:
                st.error(f"❌ Missing columns: {', '.join(missing[:5])}{'...' if len(missing) > 5 else ''}")
            elif not valid_selected:
                st.error("❌ No valid columns selected")
            else:
                out_df = df[valid_selected].copy()

                st.markdown(f"""
                <div class="success-box">
                    ✅ <strong>{tr['found_matches'].format(n=len(out_df))}</strong> ready for export
                </div>
                """, unsafe_allow_html=True)

                # Results display
                st.markdown("#### 📊 Output Preview")
                st.dataframe(out_df.head(50), use_container_width=True)

                # Prepare downloads
                excel_bytes = to_excel_bytes_styled(out_df)
                csv_bytes = out_df.to_csv(index=False).encode("utf-8")
                filename_base = "UDISE_Output" if lang == "en" else "UDISE_வெளியீடு"

                # Download buttons
                st.markdown("#### 📥 Download Options")
                dl_col1, dl_col2, dl_col3 = st.columns(3)

                with dl_col1:
                    st.download_button(
                        f"📗 {tr['download_excel']}",
                        data=excel_bytes,
                        file_name=f"{filename_base}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

                with dl_col2:
                    st.download_button(
                        f"📄 {tr['download_csv']}",
                        data=csv_bytes,
                        file_name=f"{filename_base}.csv",
                        mime="text/csv",
                        use_container_width=True
                    )

                with dl_col3:
                    copy_text = out_df.to_csv(sep="\t", index=False)
                    st.download_button(
                        f"📋 {tr['copy_output']} (TSV)",
                        data=copy_text,
                        file_name=f"{filename_base}.tsv",
                        mime="text/tab-separated-values",
                        use_container_width=True
                    )

                # Copy area for clipboard
                with st.expander("📋 Copy-Paste Area (for Excel/Sheets)"):
                    st.text_area(
                        "Select All (Ctrl+A) and Copy (Ctrl+C):",
                        copy_text,
                        height=200,
                        label_visibility="collapsed"
                    )

                st.success("💡 Excel output includes formatted headers with styling!")

# ═══════════════════════════════════════════════════════════════════════════════
# FOOTER
# ═══════════════════════════════════════════════════════════════════════════════

st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
st.markdown("""
<div style="text-align: center; color: #64748b; padding: 1rem;">
    <p>Built with ❤️ using Streamlit | UDISE Data Generator v3.0</p>
</div>
""", unsafe_allow_html=True)
